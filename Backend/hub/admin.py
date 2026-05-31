from django.contrib import admin
from django.contrib.auth.models import Group
from django import forms
from django.utils.html import format_html
from django.http import HttpResponse
from django.shortcuts import redirect
from django.core.mail import send_mail
from django.conf import settings
from django.contrib.auth.hashers import make_password
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import secrets
import string

from .models import (
    SiteContent, SiteStat, EventLog, SiteRole, User,
    Student, Guide, HubLeader, Direction,
    Laboratory, LaboratoryImage, LaboratoryGuide, LabPhoto, LaboratoryDirection,
    StudentLaboratory, StudentDirection,
    Role, LabRole, Project, ProjectLink, ProjectLaboratory,
    ProjectRole, StudentProjectRole,
    File, Achievement, FileAchievement, FileGuide, FileProject,
    Report, FileReport,
)

admin.site.unregister(Group)


def generate_password(length=8):
    return ''.join(secrets.choice(string.ascii_letters + string.digits) for _ in range(length))


# =============================================================================
# INLINES
# =============================================================================

class FileAchievementInline(admin.TabularInline):
    model = FileAchievement
    fk_name = 'achievements'
    extra = 1
    autocomplete_fields = ['file']


class FileGuideInline(admin.TabularInline):
    model = FileGuide
    fk_name = 'guide'
    extra = 1
    autocomplete_fields = ['file']


class FileProjectInline(admin.TabularInline):
    model = FileProject
    fk_name = 'project'
    extra = 1
    autocomplete_fields = ['file']


class FileReportInline(admin.TabularInline):
    model = FileReport
    fk_name = 'report'
    extra = 1
    autocomplete_fields = ['file']


class SiteStatInline(admin.TabularInline):
    model = SiteStat
    extra = 1


class ProjectLinkInline(admin.TabularInline):
    model = ProjectLink
    fk_name = 'project'
    extra = 1


class ProjectLaboratoryInline(admin.TabularInline):
    model = ProjectLaboratory
    extra = 1
    autocomplete_fields = ['laboratory']


class ProjectRoleInline(admin.TabularInline):
    model = ProjectRole
    extra = 1
    autocomplete_fields = ['role']


class LaboratoryDirectionInline(admin.TabularInline):
    model = LaboratoryDirection
    extra = 1
    autocomplete_fields = ['direction']


class GuideInline(admin.TabularInline):
    model = LaboratoryGuide
    extra = 1
    autocomplete_fields = ['guide']


# =============================================================================
# CUSTOM FILTERS
# =============================================================================

class ProjectLabFilter(admin.SimpleListFilter):
    title = 'лаборатория'
    parameter_name = 'lab'

    def lookups(self, request, model_admin):
        return Laboratory.objects.values_list('id', 'title')

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(laboratory_links__laboratory_id=self.value())
        return queryset


class StudentLabFilter(admin.SimpleListFilter):
    title = 'лаборатория'
    parameter_name = 'lab'

    def lookups(self, request, model_admin):
        return Laboratory.objects.values_list('id', 'title')

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(laboratory_links__laboratory_id=self.value())
        return queryset


class StudentDirectionFilter(admin.SimpleListFilter):
    title = 'направление'
    parameter_name = 'direction'

    def lookups(self, request, model_admin):
        return Direction.objects.values_list('id', 'title')

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(direction_links__direction_id=self.value())
        return queryset


# =============================================================================
# EXCEL EXPORT
# =============================================================================

EXCEL_HEADER_FILL = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
EXCEL_HEADER_FONT = Font(color='FFFFFF', bold=True)
EXCEL_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)


def export_to_excel(modeladmin, request, queryset):
    wb = openpyxl.Workbook()
    ws = wb.active
    opts = modeladmin.model._meta
    ws.title = opts.verbose_name_plural[:31]

    fields = modeladmin.list_display
    headers = []
    for fname in fields:
        if fname == 'action_buttons':
            headers.append('Действия')
        else:
            try:
                field = opts.get_field(fname)
                headers.append(str(field.verbose_name))
            except Exception:
                headers.append(fname)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = EXCEL_HEADER_FILL
        cell.font = EXCEL_HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
        cell.border = EXCEL_BORDER

    for row, obj in enumerate(queryset, 2):
        for col, fname in enumerate(fields, 1):
            cell = ws.cell(row=row, column=col)
            try:
                val = getattr(obj, fname)
                if callable(val):
                    val = val()
                cell.value = str(val) if val is not None else ''
            except Exception:
                cell.value = ''
            cell.border = EXCEL_BORDER

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 25

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{opts.db_table}_export.xlsx"'
    wb.save(response)
    return response


export_to_excel.short_description = 'Экспорт в Excel'


# =============================================================================
# НАСТРОЙКИ ГЛАВНОЙ СТРАНИЦЫ
# =============================================================================

@admin.register(SiteContent)
class SiteContentAdmin(admin.ModelAdmin):
    inlines = [SiteStatInline]

    def has_add_permission(self, request):
        return not SiteContent.objects.exists()

    def has_delete_permission(self, request, obj=None):
        return False


# =============================================================================
# ЖУРНАЛ СОБЫТИЙ (AuditLog)
# =============================================================================

@admin.register(EventLog)
class EventLogAdmin(admin.ModelAdmin):
    list_display = ['id', 'timestamp', 'user_login', 'action', 'entity_type', 'entity_id']
    list_filter = ['action', 'entity_type', 'timestamp']
    search_fields = ['user_login', 'entity_type']
    readonly_fields = ['id', 'user_login', 'action', 'entity_type', 'entity_id', 'timestamp', 'details']
    date_hierarchy = 'timestamp'
    actions = [export_to_excel]

    def has_add_permission(self, request):
        return False

    def has_delete_permission(self, request, obj=None):
        return False

    def has_change_permission(self, request, obj=None):
        return False


# =============================================================================
# СПРАВОЧНИКИ
# =============================================================================

@admin.register(SiteRole)
class SiteRoleAdmin(admin.ModelAdmin):
    list_display = ['id', 'title']
    search_fields = ['title']


@admin.register(Role)
class RoleAdmin(admin.ModelAdmin):
    list_display = ['id', 'title']
    search_fields = ['title']


# =============================================================================
# ПОЛЬЗОВАТЕЛИ
# =============================================================================

@admin.register(User)
class UserAdmin(admin.ModelAdmin):
    list_display = ['id', 'login', 'site_role', 'laboratory']
    list_filter = ['site_role', 'laboratory']
    search_fields = ['login']
    raw_id_fields = ['site_role', 'laboratory']
    actions = [export_to_excel]

    def save_model(self, request, obj, form, change):
        if not change:
            obj.password = make_password(obj.password) if obj.password and not obj.password.startswith('pbkdf2_') else make_password(generate_password())
        elif obj.password and not obj.password.startswith('pbkdf2_'):
            obj.password = make_password(obj.password)
        super().save_model(request, obj, form, change)


class StudentLaboratoryInline(admin.TabularInline):
    model = StudentLaboratory
    fk_name = 'student'
    extra = 1
    autocomplete_fields = ['laboratory']
    verbose_name = 'Лаборатория'
    verbose_name_plural = 'Лаборатории'


@admin.register(Student)
class StudentAdmin(admin.ModelAdmin):
    inlines = [StudentLaboratoryInline]
    list_display = ['id', 'surname', 'name', 'study_group', 'email', 'telegram_nickname', 'experience']
    list_filter = ['study_group', 'university_city', 'experience', StudentLabFilter, StudentDirectionFilter]
    search_fields = ['surname', 'name', 'email', 'telegram_nickname', 'study_group']
    actions = [export_to_excel]

    def get_queryset(self, request):
        return super().get_queryset(request).prefetch_related(
            'laboratory_links', 'direction_links', 'project_roles'
        )


@admin.register(Guide)
class GuideAdmin(admin.ModelAdmin):
    inlines = [FileGuideInline]
    list_display = ['id', 'surname', 'name', 'patronymic', 'position']
    search_fields = ['surname', 'name', 'patronymic', 'position']
    actions = [export_to_excel]

    def full_name(self, obj):
        return f'{obj.surname} {obj.name} {obj.patronymic or ""}'.strip()
    full_name.short_description = 'ФИО'


@admin.register(HubLeader)
class HubLeaderAdmin(admin.ModelAdmin):
    list_display = ['id', 'full_name', 'user', 'position', 'degree', 'phone', 'email', 'is_active']
    list_filter = ['is_active']
    search_fields = ['full_name', 'user__login', 'position', 'degree', 'phone', 'email']
    raw_id_fields = ['user']


# =============================================================================
# НАПРАВЛЕНИЯ
# =============================================================================

@admin.register(Direction)
class DirectionAdmin(admin.ModelAdmin):
    list_display = ['id', 'title', 'labs_count']
    search_fields = ['title']

    def labs_count(self, obj):
        return obj.laboratory_links.count()
    labs_count.short_description = 'Количество лабораторий'


# =============================================================================
# ЛАБОРАТОРИИ
# =============================================================================

class LaboratoryImageInline(admin.TabularInline):
    model = LaboratoryImage
    extra = 0
    max_num = 1
    autocomplete_fields = ['lab_photo']


@admin.register(LabPhoto)
class LabPhotoAdmin(admin.ModelAdmin):
    list_display = ['id', 'card_image_tag', 'card_image', 'lab_image_tag', 'lab_image']
    search_fields = ['card_image']

    def card_image_tag(self, obj):
        if obj.card_image:
            return format_html('<img src="{}" style="max-height:80px;max-width:120px" />', obj.card_image.url)
        return ''
    card_image_tag.short_description = 'Превью (карточка)'

    def lab_image_tag(self, obj):
        if obj.lab_image:
            return format_html('<img src="{}" style="max-height:80px;max-width:120px" />', obj.lab_image.url)
        return ''
    lab_image_tag.short_description = 'Превью (колба)'


@admin.register(Laboratory)
class LaboratoryAdmin(admin.ModelAdmin):
    inlines = [LaboratoryDirectionInline, GuideInline, LaboratoryImageInline]
    list_display = ['id', 'title', 'slug', 'active', 'link']
    list_filter = ['active']
    search_fields = ['title', 'slug']
    actions = [export_to_excel]


@admin.register(LaboratoryDirection)
class LaboratoryDirectionAdmin(admin.ModelAdmin):
    list_display = ['id', 'laboratory', 'direction', 'link']
    list_filter = ['laboratory', 'direction']
    autocomplete_fields = ['laboratory', 'direction']


@admin.register(StudentLaboratory)
class StudentLaboratoryAdmin(admin.ModelAdmin):
    list_display = ['id', 'student', 'laboratory']
    list_filter = ['laboratory']
    search_fields = ['student__surname', 'student__name']
    autocomplete_fields = ['student', 'laboratory']


@admin.register(StudentDirection)
class StudentDirectionAdmin(admin.ModelAdmin):
    list_display = ['id', 'student', 'direction']
    list_filter = ['direction']
    search_fields = ['student__surname', 'student__name']
    autocomplete_fields = ['student', 'direction']


# =============================================================================
# РОЛИ ЛАБОРАТОРИЙ
# =============================================================================

@admin.register(LabRole)
class LabRoleAdmin(admin.ModelAdmin):
    list_display = ['id', 'title', 'laboratory']
    list_filter = ['laboratory']
    search_fields = ['title']
    autocomplete_fields = ['laboratory']


# =============================================================================
# ПРОЕКТЫ
# =============================================================================

@admin.register(Project)
class ProjectAdmin(admin.ModelAdmin):
    inlines = [ProjectLinkInline, ProjectLaboratoryInline, ProjectRoleInline, FileProjectInline]
    list_display = ['id', 'title', 'description', 'need_report']
    list_filter = ['need_report', ProjectLabFilter]
    search_fields = ['title', 'description', 'goal']
    actions = [export_to_excel]
    change_form_template = 'admin/hub/project/change_form.html'

    def render_change_form(self, request, context, *args, **kwargs):
        instance = context.get('original')
        if instance:
            context['project_participants'] = StudentProjectRole.objects.filter(
                project_role__project=instance
            ).select_related('student', 'project_role__role')
            lab_ids = instance.laboratory_links.values_list('laboratory_id', flat=True)
            context['available_students'] = Student.objects.filter(
                laboratory_links__laboratory_id__in=lab_ids
            ).distinct().order_by('surname', 'name')
            context['project_roles_qs'] = ProjectRole.objects.filter(
                project=instance
            ).select_related('role')
        return super().render_change_form(request, context, *args, **kwargs)

    def response_change(self, request, obj):
        if 'delete_participant' in request.POST:
            participant_id = request.POST.get('delete_participant')
            if participant_id:
                StudentProjectRole.objects.filter(id=participant_id).delete()
            return redirect(request.path)

        if 'add_participant' in request.POST:
            student_id = request.POST.get('new_participant_student')
            project_role_id = request.POST.get('new_participant_role')
            if student_id and project_role_id:
                lab_ids = obj.laboratory_links.values_list('laboratory_id', flat=True)
                if StudentLaboratory.objects.filter(
                    student_id=student_id, laboratory_id__in=lab_ids
                ).exists():
                    StudentProjectRole.objects.get_or_create(
                        student_id=student_id,
                        project_role_id=project_role_id,
                        defaults={'present': True},
                    )
            return redirect(request.path)

        return super().response_change(request, obj)


@admin.register(ProjectLink)
class ProjectLinkAdmin(admin.ModelAdmin):
    list_display = ['id', 'title', 'url', 'project']
    list_filter = ['project']
    search_fields = ['title', 'url']
    autocomplete_fields = ['project']


@admin.register(ProjectLaboratory)
class ProjectLaboratoryAdmin(admin.ModelAdmin):
    list_display = ['id', 'project', 'laboratory']
    list_filter = ['laboratory', 'project']
    autocomplete_fields = ['project', 'laboratory']


@admin.register(ProjectRole)
class ProjectRoleAdmin(admin.ModelAdmin):
    list_display = ['id', 'project', 'role']
    list_filter = ['project']
    autocomplete_fields = ['project', 'role']
    search_fields = ['project__title', 'role__title']


@admin.register(StudentProjectRole)
class StudentProjectRoleAdmin(admin.ModelAdmin):
    list_display = ['id', 'student', 'project_role', 'present']
    list_filter = ['present', 'project_role__project']
    search_fields = ['student__surname', 'student__name']
    autocomplete_fields = ['student', 'project_role']
    actions = [export_to_excel]


# =============================================================================
# ФАЙЛЫ
# =============================================================================

@admin.register(File)
class FileAdmin(admin.ModelAdmin):
    list_display = ['id', 'file_link']
    search_fields = ['link']

    def file_link(self, obj):
        return format_html('<a href="{}" target="_blank">{}</a>', obj.link, obj.link[:80])
    file_link.short_description = 'Ссылка'


# =============================================================================
# ДОСТИЖЕНИЯ
# =============================================================================

@admin.register(Achievement)
class AchievementAdmin(admin.ModelAdmin):
    inlines = [FileAchievementInline]
    list_display = ['id', 'title', 'laboratory', 'project']
    list_filter = ['laboratory', 'project']
    search_fields = ['title', 'description']
    autocomplete_fields = ['laboratory', 'project']
    actions = [export_to_excel]


# =============================================================================
# ОТЧЁТЫ
# =============================================================================

@admin.register(Report)
class ReportAdmin(admin.ModelAdmin):
    inlines = [FileReportInline]
    list_display = ['id', 'project', 'laboratory', 'date_time', 'confirmation']
    list_filter = ['confirmation', 'laboratory', 'project']
    search_fields = ['report_text']
    readonly_fields = ['date_time']
    autocomplete_fields = ['laboratory', 'project', 'file_report']
    actions = [export_to_excel]


# =============================================================================
# ПРОМЕЖУТОЧНЫЕ ТАБЛИЦЫ
# =============================================================================

@admin.register(FileAchievement)
class FileAchievementAdmin(admin.ModelAdmin):
    list_display = ['id', 'file', 'achievements']
    autocomplete_fields = ['file', 'achievements']


@admin.register(FileGuide)
class FileGuideAdmin(admin.ModelAdmin):
    list_display = ['id', 'file', 'guide']
    autocomplete_fields = ['file', 'guide']


@admin.register(FileProject)
class FileProjectAdmin(admin.ModelAdmin):
    list_display = ['id', 'file', 'project']
    autocomplete_fields = ['file', 'project']


@admin.register(FileReport)
class FileReportAdmin(admin.ModelAdmin):
    list_display = ['id', 'file', 'report']
    autocomplete_fields = ['file', 'report']
