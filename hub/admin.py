from django.contrib import admin
from django.contrib.admin import ModelAdmin
from django.db import models
from django.utils.html import format_html
from django.http import HttpResponse
from django.utils import timezone
from django.core.mail import send_mail
from django.conf import settings
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from .models import (
    SiteRole, Status, Type,
    User, Student, Guide,
    Direction, Laboratory, LaboratoryDirection, LaboratoryLeader,
    StudentLaboratory, StudentDirection,
    Role, Project, ProjectLaboratory, ProjectRole, StudentProjectRole,
    File, Achievement, FileAchievement, FileGuide, FileProject,
    Report, FileReport, EventLog, HubLeader
)


# =============================================================================
# УТИЛИТЫ
# =============================================================================

def export_to_excel(modeladmin, request, queryset):
    wb = Workbook()
    ws = wb.active
    ws.title = modeladmin.model._meta.verbose_name_plural
    headers = ['ID'] + [f.verbose_name for f in modeladmin.model._meta.fields if f.name != 'id']
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row_num, obj in enumerate(queryset, 2):
        for col_num, field in enumerate(modeladmin.model._meta.fields, 1):
            val = getattr(obj, field.name, '')
            if hasattr(val, 'strftime'):
                val = val.strftime('%d.%m.%Y %H:%M')
            ws.cell(row=row_num, column=col_num, value=str(val) if val is not None else '')
    for col_num in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 25
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"{modeladmin.model._meta.verbose_name_plural}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
export_to_excel.short_description = "Экспортировать выбранные в Excel"


# =============================================================================
# СПРАВОЧНИКИ
# =============================================================================

@admin.register(SiteRole)
class SiteRoleAdmin(ModelAdmin):
    list_display = ['id', 'title']
    search_fields = ['title']


@admin.register(Status)
class StatusAdmin(ModelAdmin):
    list_display = ['id', 'title']
    search_fields = ['title']


@admin.register(Type)
class TypeAdmin(ModelAdmin):
    list_display = ['id', 'title']
    search_fields = ['title']


# =============================================================================
# ПОЛЬЗОВАТЕЛИ
# =============================================================================

@admin.register(User)
class UserAdmin(ModelAdmin):
    list_display = ['id', 'login', 'site_role', 'laboratory']
    list_filter = ['site_role', 'laboratory']
    search_fields = ['login']
    fieldsets = (
        (None, {'fields': ('login', 'password')}),
        ('Роль и лаборатория', {'fields': ('site_role', 'laboratory')}),
    )
    def save_model(self, request, obj, form, change):
        if 'password' in form.changed_data:
            from django.contrib.auth.hashers import make_password
            obj.password = make_password(form.cleaned_data['password'])
        super().save_model(request, obj, form, change)


@admin.register(Student)
class StudentAdmin(ModelAdmin):
    list_display = [
        'id', 'surname', 'name', 'patronymic', 'study_group',
        'email', 'phone_number', 'telegram_nickname', 'experience_badge'
    ]
    list_filter = ['study_group', 'experience', 'university_city']
    search_fields = ['surname', 'name', 'email', 'telegram_nickname', 'study_group']
    fieldsets = (
        ('ФИО', {'fields': ('surname', 'name', 'patronymic')}),
        ('Учёба', {'fields': ('study_group', 'university_city')}),
        ('Контакты', {'fields': ('phone_number', 'email', 'telegram_nickname', 'telegram_id')}),
        ('Опыт и пожелания', {'fields': ('experience', 'wishes')}),
        ('Метавселенная', {'fields': ('metaverse_account_link', 'task_board')}),
    )

    def experience_badge(self, obj):
        if obj.experience == 'Да':
            return format_html('<span style="color: green; font-weight: bold;">{}</span>', obj.experience or '-')
        elif obj.experience == 'Немного':
            return format_html('<span style="color: orange;">{}</span>', obj.experience or '-')
        elif obj.experience == 'Нет':
            return format_html('<span style="color: gray;">{}</span>', obj.experience or '-')
        return '-'
    experience_badge.short_description = 'Опыт'

    def has_add_permission(self, request):
        return False


@admin.register(Guide)
class GuideAdmin(ModelAdmin):
    list_display = ['id', 'surname', 'name', 'patronymic', 'laboratory', 'full_fio']
    list_filter = ['laboratory']
    search_fields = ['surname', 'name', 'patronymic']

    def full_fio(self, obj):
        return obj.__str__()
    full_fio.short_description = 'ФИО'


# =============================================================================
# ЛАБОРАТОРИИ И НАПРАВЛЕНИЯ
# =============================================================================

@admin.register(Direction)
class DirectionAdmin(ModelAdmin):
    list_display = ['id', 'title', 'labs_count']
    search_fields = ['title']

    def labs_count(self, obj):
        return obj.laboratory_links.count()
    labs_count.short_description = 'Лабораторий'


@admin.register(Laboratory)
class LaboratoryAdmin(ModelAdmin):
    list_display = ['id', 'title', 'active', 'short_desc_preview', 'link', 'leaders_count', 'students_count', 'directions_list']
    list_filter = ['active']
    search_fields = ['title', 'short_description']
    actions = [export_to_excel]

    fieldsets = (
        (None, {'fields': ('title', 'link', 'active')}),
        ('Описание и фото', {'fields': ('short_description', 'images')}),
    )

    def short_desc_preview(self, obj):
        if obj.short_description:
            return obj.short_description[:60] + '...' if len(obj.short_description) > 60 else obj.short_description
        return '-'
    short_desc_preview.short_description = 'Краткое описание'

    def directions_list(self, obj):
        return ', '.join([ld.direction.title for ld in obj.direction_links.all()]) or '-'
    directions_list.short_description = 'Направления'

    def leaders_count(self, obj):
        return obj.leaders.count()
    leaders_count.short_description = 'Лидеров'

    def students_count(self, obj):
        return obj.student_links.filter(laboratory__isnull=False).count()
    students_count.short_description = 'Студентов'


@admin.register(LaboratoryDirection)
class LaboratoryDirectionAdmin(ModelAdmin):
    list_display = ['id', 'laboratory', 'direction', 'link']
    list_filter = ['laboratory', 'direction']
    search_fields = ['laboratory__title', 'direction__title']


@admin.register(LaboratoryLeader)
class LaboratoryLeaderAdmin(ModelAdmin):
    list_display = ['id', 'student_info', 'laboratory', 'student_email', 'student_group']
    list_filter = ['laboratory']
    search_fields = ['student__surname', 'student__name', 'laboratory__title']
    actions = [export_to_excel, 'assign_as_manager', 'assign_as_hub_leader']

    def student_info(self, obj):
        return obj.student.full_name
    student_info.short_description = 'Студент'

    def student_email(self, obj):
        return obj.student.email
    student_email.short_description = 'Email'

    def student_group(self, obj):
        return obj.student.study_group
    student_group.short_description = 'Группа'

    def assign_as_manager(self, request, queryset):
        """Назначить менеджером + отправить email с паролем"""
        import random, string, secrets
        count = 0
        for leader in queryset:
            student = leader.student
            if student.email:
                password = ''.join(secrets.choice(string.ascii_letters + string.digits) for _ in range(8))
                # Генерация аккаунта в user-таблице (или обновление)
                user, created = User.objects.get_or_create(login=student.email.split('@')[0])
                from django.contrib.auth.hashers import make_password
                user.password = make_password(password)
                user.laboratory = leader.laboratory
                user.save()

                try:
                    send_mail(
                        f'Вы назначены менеджером проекта — {leader.laboratory.title}',
                        f'''Здравствуйте, {student.full_name}!

Вы назначены менеджером лаборатории "{leader.laboratory.title}".

Ваш логин: {user.login}
Ваш пароль: {password}

Войдите в админ-панель и измените пароль.''',
                        settings.DEFAULT_FROM_EMAIL,
                        [student.email],
                        fail_silently=True
                    )
                except Exception:
                    pass
                count += 1
        self.message_user(request, f'{count} студентов назначено менеджерами. Email отправлены.')
    assign_as_manager.short_description = 'Назначить менеджером проекта'

    def assign_as_hub_leader(self, request, queryset):
        """Назначить руководителем хаба"""
        count = 0
        for leader in queryset:
            student = leader.student
            obj, created = HubLeader.objects.get_or_create(
                student=student,
                defaults={'position': 'Руководитель лаборатории', 'is_active': True}
            )
            if not created:
                obj.is_active = True
                obj.save()
            count += 1
        self.message_user(request, f'{count} студентов назначено руководителями хаба.')
    assign_as_hub_leader.short_description = 'Назначить руководителем хаба'


@admin.register(StudentLaboratory)
class StudentLaboratoryAdmin(ModelAdmin):
    list_display = ['id', 'student', 'laboratory', 'student_group']
    list_filter = ['laboratory']
    search_fields = ['student__surname', 'student__name']

    def student_group(self, obj):
        return obj.student.study_group
    student_group.short_description = 'Группа'


@admin.register(StudentDirection)
class StudentDirectionAdmin(ModelAdmin):
    list_display = ['id', 'student', 'direction', 'student_group']
    list_filter = ['direction']
    search_fields = ['student__surname', 'student__name', 'direction__title']

    def student_group(self, obj):
        return obj.student.study_group
    student_group.short_description = 'Группа'


# =============================================================================
# ПРОЕКТЫ И РОЛИ
# =============================================================================

@admin.register(Role)
class RoleAdmin(ModelAdmin):
    list_display = ['id', 'title']
    search_fields = ['title']
    actions = [export_to_excel]


@admin.register(Project)
class ProjectAdmin(ModelAdmin):
    list_display = ['id', 'title', 'goal_preview', 'description_preview', 'need_report', 'labs_list', 'participants_count']
    list_filter = ['need_report']
    search_fields = ['title', 'description', 'goal']
    actions = [export_to_excel]

    fieldsets = (
        (None, {'fields': ('title', 'need_report')}),
        ('Описание', {'fields': ('description', 'goal')}),
    )

    def goal_preview(self, obj):
        if obj.goal:
            return obj.goal[:60] + '...' if len(obj.goal) > 60 else obj.goal
        return '-'
    goal_preview.short_description = 'Цель'

    def description_preview(self, obj):
        if obj.description:
            return obj.description[:80] + '...' if len(obj.description) > 80 else obj.description
        return '-'
    description_preview.short_description = 'Описание'

    def labs_list(self, obj):
        labs = [pl.laboratory.title for pl in obj.laboratory_links.all()]
        return ', '.join(labs) if labs else '-'
    labs_list.short_description = 'Лаборатории'

    def participants_count(self, obj):
        return obj.studentprojectrole_set.filter(present=True).count()
    participants_count.short_description = 'Участников'


@admin.register(ProjectLaboratory)
class ProjectLaboratoryAdmin(ModelAdmin):
    list_display = ['id', 'project', 'laboratory']
    list_filter = ['laboratory']
    search_fields = ['project__title', 'laboratory__title']


@admin.register(ProjectRole)
class ProjectRoleAdmin(ModelAdmin):
    list_display = ['id', 'project', 'role', 'assigned_count']
    list_filter = ['role']
    search_fields = ['project__title', 'role__title']

    def assigned_count(self, obj):
        return obj.assigned_students.filter(present=True).count()
    assigned_count.short_description = 'Занято'


@admin.register(StudentProjectRole)
class StudentProjectRoleAdmin(ModelAdmin):
    list_display = ['id', 'student', 'project', 'role', 'status_display']
    list_filter = ['present', 'project_role__role']
    search_fields = ['student__surname', 'student__name', 'project_role__project__title']
    actions = [export_to_excel, 'mark_as_left', 'send_manager_email']

    fieldsets = (
        (None, {'fields': ('student', 'project_role', 'present')}),
    )

    def project(self, obj):
        return obj.project_role.project.title
    project.short_description = 'Проект'

    def role(self, obj):
        return obj.project_role.role.title
    role.short_description = 'Роль'

    def status_display(self, obj):
        if obj.present:
            return format_html('<span style="color: green;">Активен</span>')
        return format_html('<span style="color: red;">Покинул</span>')
    status_display.short_description = 'Статус'

    def mark_as_left(self, request, queryset):
        count = queryset.filter(present=True).update(present=False)
        self.message_user(request, f'{count} участников помечено как покинувших проект.')
    mark_as_left.short_description = "Пометить как покинувших проект"

    def send_manager_email(self, request, queryset):
        """Отправить email с паролем выбранным менеджерам"""
        import random, string, secrets
        count = 0
        for participant in queryset.filter(present=True):
            student = participant.student
            if student.email:
                password = ''.join(secrets.choice(string.ascii_letters + string.digits) for _ in range(8))
                user, created = User.objects.get_or_create(login=student.email.split('@')[0])
                from django.contrib.auth.hashers import make_password
                user.password = make_password(password)
                user.save()

                try:
                    send_mail(
                        f'Вы назначены менеджером — {participant.project_role.project.title}',
                        f'''Здравствуйте, {student.full_name}!

Вам назначена роль менеджера в проекте "{participant.project_role.project.title}".

Ваш логин: {user.login}
Ваш пароль: {password}

Войдите в админ-панель.''',
                        settings.DEFAULT_FROM_EMAIL,
                        [student.email],
                        fail_silently=True
                    )
                    count += 1
                except Exception:
                    pass
        self.message_user(request, f'{count} email отправлено менеджерам.')
    send_manager_email.short_description = 'Отправить пароль менеджерам'


# =============================================================================
# ДОСТИЖЕНИЯ И ФАЙЛЫ
# =============================================================================

@admin.register(File)
class FileAdmin(ModelAdmin):
    list_display = ['id', 'link_preview']
    search_fields = ['link']

    def link_preview(self, obj):
        name = obj.link.split('/')[-1] if obj.link else str(obj.id)
        return name[:60]
    link_preview.short_description = 'Файл'


@admin.register(Achievement)
class AchievementAdmin(ModelAdmin):
    list_display = ['id', 'title', 'laboratory', 'project', 'text_limited_preview', 'created']
    list_filter = ['laboratory']
    search_fields = ['title', 'text', 'text_limited']
    actions = [export_to_excel]

    fieldsets = (
        (None, {'fields': ('title', 'laboratory', 'project')}),
        ('Описание (до 350 символов)', {'fields': ('text', 'text_limited', 'link')}),
        ('Файл', {'fields': ('file_achievements',)}),
    )

    def text_limited_preview(self, obj):
        text = obj.text_limited or obj.text or ''
        if text:
            return text[:60] + '...' if len(text) > 60 else text
        return '-'
    text_limited_preview.short_description = 'Описание (350)'

    def created(self, obj):
        return obj.id  # сортировка по id (он по порядку)
    created.short_description = 'ID'


@admin.register(FileAchievement)
class FileAchievementAdmin(ModelAdmin):
    list_display = ['id', 'file', 'achievements']
    list_filter = ['achievements']


@admin.register(FileGuide)
class FileGuideAdmin(ModelAdmin):
    list_display = ['id', 'file', 'guide']
    list_filter = ['guide']


@admin.register(FileProject)
class FileProjectAdmin(ModelAdmin):
    list_display = ['id', 'file', 'project']
    list_filter = ['project']


# =============================================================================
# ОТЧЁТЫ
# =============================================================================

@admin.register(Report)
class ReportAdmin(ModelAdmin):
    list_display = ['id', 'project', 'laboratory', 'date_time', 'confirmation_display']
    list_filter = ['confirmation', 'laboratory']
    search_fields = ['project__title', 'report_text']
    actions = [export_to_excel, 'confirm_reports']

    def confirmation_display(self, obj):
        if obj.confirmation:
            return format_html('<span style="color: green;">✓ Подтверждён</span>')
        return format_html('<span style="color: orange;">Ожидает</span>')
    confirmation_display.short_description = 'Статус'

    def confirm_reports(self, request, queryset):
        count = queryset.filter(confirmation=False).update(confirmation=True)
        self.message_user(request, f'{count} отчётов подтверждено.')
    confirm_reports.short_description = 'Подтвердить выбранные'


@admin.register(FileReport)
class FileReportAdmin(ModelAdmin):
    list_display = ['id', 'file', 'report']
    list_filter = ['report__laboratory']


# =============================================================================
# РУКОВОДИТЕЛИ ХАБА
# =============================================================================

@admin.register(HubLeader)
class HubLeaderAdmin(ModelAdmin):
    list_display = ['id', 'student', 'position', 'phone', 'is_active', 'created_at', 'labs_count']
    list_filter = ['is_active']
    search_fields = ['student__surname', 'student__name', 'student__email', 'position']
    actions = [export_to_excel, 'activate_leaders', 'deactivate_leaders']

    fieldsets = (
        (None, {'fields': ('student', 'position', 'phone', 'is_active')}),
        ('Дата', {'fields': ('created_at',), 'classes': ('collapse',)}),
    )

    readonly_fields = ['created_at']

    def labs_count(self, obj):
        labs = list(obj.student.laboratory_links.filter(laboratory__isnull=False).values_list('laboratory__title', flat=True))
        return ', '.join(labs) if labs else '-'
    labs_count.short_description = 'Лаборатории'

    def activate_leaders(self, request, queryset):
        queryset.update(is_active=True)
        self.message_user(request, f'{queryset.count()} руководителей активировано.')
    activate_leaders.short_description = 'Активировать руководителей'

    def deactivate_leaders(self, request, queryset):
        queryset.update(is_active=False)
        self.message_user(request, f'{queryset.count()} руководителей деактивировано.')
    deactivate_leaders.short_description = 'Деактивировать руководителей'


# =============================================================================
# ЖУРНАЛ СОБЫТИЙ
# =============================================================================

@admin.register(EventLog)
class EventLogAdmin(ModelAdmin):
    list_display = ['id', 'timestamp', 'user_login', 'action', 'entity_type', 'entity_id', 'details_preview']
    list_filter = ['action', 'entity_type', 'timestamp']
    search_fields = ['user_login', 'entity_type']
    actions = [export_to_excel]

    def details_preview(self, obj):
        if obj.details:
            return obj.details[:60] if len(obj.details) > 60 else obj.details
        return '-'
    details_preview.short_description = 'Детали'

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False
